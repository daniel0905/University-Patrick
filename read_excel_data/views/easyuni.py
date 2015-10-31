import os
from os import walk
from openpyxl import load_workbook
from django.http import HttpResponse

from university.models import *
DATA_FOLDER_PATH = "C:\\Users\\dend\\Desktop\\Scrapy\\Easyuni-data\\Universitys"

def import_data_easyuni(request):
	for folder in range(1, 99):
		page_folder_path = os.path.join(DATA_FOLDER_PATH, str(folder))
		for (dir_path, dir_names, file_names) in walk(page_folder_path):
			for file_name in file_names:
				uni_file_path = os.path.join(page_folder_path, file_name)
				wb = load_workbook(uni_file_path)
				university_sheet = wb["university"]

				# Import data to Database
				university = import_university_information(university_sheet)
				if university:
					course_sheet = wb["courses"]
					video_sheet = wb["video"]
					linkedin_sheet = wb["linkedin"]

					import_reason(university, university_sheet)
					import_course(university, course_sheet)
					import_video(university, video_sheet)
					import_linkedin(university, linkedin_sheet)
			break
	return HttpResponse("Successful")

def import_university_information(university_sheet):
	# FIXME: Change University Information in here
	name_university_sheet = university_sheet.cell("A2").value
	name_university = name_university_sheet[0:name_university_sheet.find(",")]

	university_info = dict()
	university_info["about"] = university_sheet.cell("K2").value
	university_info["history"] = university_sheet.cell("L2").value
	university_info["year_founded"] = university_sheet.cell("C2").value

	university_info["student_sum"] = university_sheet.cell("D2").value
	university_info["male_female_ratio"] = university_sheet.cell("H2").value
	university_info["local_international_ratio"] = university_sheet.cell("I2").value
	university_info["alumni"] = university_sheet.cell("P2").value

	university_info["campus"] = university_sheet.cell("M2").value
	university_info["accommodation"] = university_sheet.cell("O2").value
	university_info["location_description"] = university_sheet.cell("N2").value

	university_info["facebook_link"] = university_sheet.cell("G2").value
	university_info["twitter_link"] = university_sheet.cell("J2").value
	university_info["linkedin_link"] = university_sheet.cell("T2").value

	university_info["awards"] = university_sheet.cell("Q2").value
	university_info["additional_info"] = university_sheet.cell("R2").value
	university_info["disclaimer"] = university_sheet.cell("S2").value

	university = University.create_new_university(name=name_university, **university_info)
	return university

def import_reason(university, university_sheet):
	# FIXME: Change University Reason in here
	for x in range(1, 5):
		if university_sheet.columns[20][x].value or university_sheet.columns[21][x].value:
			reason_info = dict()
			reason_info["university"] = university
			reason_info["title"] = university_sheet.columns[20][x].value
			reason_info["content"] = university_sheet.columns[21][x].value
			Reason.create_new_reason(**reason_info)
		else:
			break

def import_course(university, course_sheet):
	# FIXME: Change University Course in here
	courses = course_sheet.columns[0]
	if len(courses) > 1:
		for x in range(1, len(courses)):
			course_info = dict()
			course_info["university"] = university
			course_info["name"] = course_sheet.columns[1][x].value
			course_info["level"] = course_sheet.columns[3][x].value
			course_info["awarded_by"] = course_sheet.columns[4][x].value
			course_info["duration"] = course_sheet.columns[5][x].value
			course_info["study_mode"] = course_sheet.columns[6][x].value

			course_info["about_course"] = course_sheet.columns[13][x].value
			course_info["entry_requirements"] = course_sheet.columns[14][x].value
			course_info["subjects"] = course_sheet.columns[15][x].value

			course_info["tuition_fee_local_per_year"] = course_sheet.columns[8][x].value
			course_info["tuition_fee_local_entire"] = course_sheet.columns[9][x].value
			course_info["tuition_fee_inter_per_year"] = course_sheet.columns[10][x].value
			course_info["tuition_fee_inter_entire"] = course_sheet.columns[11][x].value
			course_info["other_costs"] = course_sheet.columns[12][x].value

			Course.create_new_course(**course_info)

def import_video(university, video_sheet):
	# FIXME: Change University Video in here
	videos = video_sheet.columns[0]
	if len(videos) > 1:
		for x in range(1, len(videos)):
			video_info = dict()
			video_info["university"] = university
			video_info["title"] = video_sheet.columns[0][x].value
			video_info["url"] = video_sheet.columns[1][x].value

			Video.create_new_video(**video_info)


def import_linkedin(university, linkedin_sheet):
	# FIXME: Change University Linkedin in here
	linkedin = linkedin_sheet.columns[0]
	if len(linkedin) > 1:
		for x in range(1, len(linkedin)):
			linkedin_info = dict()
			linkedin_info["university"] = university
			linkedin_info["group_linkedin"] = linkedin_sheet.columns[0][x].value
			linkedin_info["title"] = linkedin_sheet.columns[1][x].value
			linkedin_info["count"] = linkedin_sheet.columns[2][x].value

			LinkedIn.create_new_linkedin(**linkedin_info)



