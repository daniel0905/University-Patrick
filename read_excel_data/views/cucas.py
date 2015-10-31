import os
import logging
from os import walk
from openpyxl import load_workbook
from django.http import HttpResponse

from university.models import *
DATA_FOLDER_PATH = "C:\\Users\\dend\\Desktop\\Cucas\\Cucas\\Data"

def import_data_cucas(request):
	folder_data_path = list(walk(DATA_FOLDER_PATH))
	for index, folder in enumerate(folder_data_path):
		if index > 0:
			folder_university_path = folder[0]
			for (dir_path, dir_names, file_name) in walk(folder_university_path):
				uni_file_path = os.path.join(folder_university_path, str(file_name[0]))
				wb = load_workbook(uni_file_path)
				university_sheet = wb["University"]

				logging.warning(file_name[0])
				university = import_university_information(university_sheet)
				if university:
					import_reason(university, university_sheet)
					import_course(university, university_sheet)

	return HttpResponse("Successful")

def import_university_information(university_sheet):
	# FIXME: Change University Information in here
	name_university = university_sheet.cell("A2").value
	university_info = dict()
	university_info["location"] = university_sheet.cell("C2").value
	university_info["living_expense"] = university_sheet.cell("E2").value
	university = University.create_new_university(name=name_university, **university_info)
	return university

def import_reason(university, university_sheet):
	reason_info = dict()
	reason_info["university"] = university
	reason_info["content"] = university_sheet.cell("D2").value
	Reason.create_new_reason(**reason_info)

def import_course(university, university_sheet):
	# FIXME: Change University Course in here
	courses = university_sheet.columns[0]
	if len(courses) > 1:
		for x in range(1, len(courses)):
			logging.warning(x)
			course_info = dict()
			course_info["university"] = university
			course_info["name"] = university_sheet.columns[5][x].value
			course_info["duration"] = university_sheet.columns[8][x].value
			course_info["about_course"] = university_sheet.columns[19][x].value

			course_info["qualification_awarded"] = university_sheet.columns[6][x].value
			course_info["language"] = university_sheet.columns[7][x].value
			course_info["starting_date"] = university_sheet.columns[10][x].value
			course_info["application_deadline"] = university_sheet.columns[11][x].value
			course_info["application_fee"] = university_sheet.columns[12][x].value
			course_info["academic_requirement"] = university_sheet.columns[13][x].value
			course_info["enrollment_quota"] = university_sheet.columns[14][x].value
			course_info["affiliated_hospitals"] = university_sheet.columns[15][x].value
			course_info["english_requirement"] = university_sheet.columns[16][x].value
			course_info["admission_difficulty"] = university_sheet.columns[17][x].value
			course_info["hsk_requirement"] = university_sheet.columns[18][x].value
			course_info["application_materials"] = university_sheet.columns[20][x].value

			course_info["accommodation_cost"] = university_sheet.columns[22][x].value
			course_info["tuition_fee_living"] = university_sheet.columns[23][x].value
			course_info["other_costs"] = university_sheet.columns[24][x].value
			course_info["tuition_fee_local_entire"] = university_sheet.columns[25][x].value
			course_info["tuition_fee_local_per_year"] = university_sheet.columns[11][x].value

			Course.create_new_course(**course_info)




